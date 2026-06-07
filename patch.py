import re

with open('backend/engines/export_engine.py', 'r', encoding='utf-8') as f:
    code = f.read()

new_func = r'''    def to_excel_advanced(data: list, project_name: str = "METRAI EXPERT") -> bytes:
        """
        Génère un fichier Excel Avancé avec 2 onglets:
        1. Détails de Fabrication (Regroupés par famille)
        2. Synthèse Globale (Dashboard)
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from itertools import groupby
        from datetime import datetime
        import re

        wb = Workbook()
        
        # --- ONGLET 1: Détails de Fabrication ---
        ws_details = wb.active
        ws_details.title = "Détails de Fabrication"
        
        # Styles
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        thick_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
        
        # Cartouche (En-tête professionnel)
        ws_details.merge_cells('B2:F3')
        title_cell = ws_details.cell(row=2, column=2, value="NOMENCLATURE ET LISTE DE DÉBIT - CHARPENTE MÉTALLIQUE")
        title_cell.font = Font(name="Calibri", bold=True, size=16, color="1F4E78")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.border = thick_border
        
        ws_details.merge_cells('G2:I2')
        ws_details.cell(row=2, column=7, value=f"Projet : {project_name}").font = Font(bold=True)
        ws_details.cell(row=2, column=7).border = thick_border
        
        ws_details.merge_cells('G3:I3')
        ws_details.cell(row=3, column=7, value=f"Date : {datetime.now().strftime('%d/%m/%Y')}").font = Font(italic=True)
        ws_details.cell(row=3, column=7).border = thick_border
        
        headers = ["Repère", "Nomenclature", "Profilé", "Nuance", "Long (mm)", "Quantité", "Poids Unit (Kg)", "Poids Tot (Kg)", "Observation"]
        for col_num, header in enumerate(headers, 1):
            cell = ws_details.cell(row=6, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            
        # Groupement par famille
        def get_famille(designation):
            des = str(designation).upper()
            if re.match(r'^(IPE|HEA|HEB|HEM|UPN|UPE|INP)', des): return "PROFILÉS STANDARDS (I, H, U)"
            if re.match(r'^(L|2L|CORNI)', des): return "CORNIÈRES (L)"
            if re.match(r'^(PL|TN|PLAT|TÔLE)', des): return "TÔLES ET PLATINES"
            if re.match(r'^(TU|CHS|RHS|SHS)', des): return "TUBES"
            if re.match(r'^(BOU|M\d+)', des): return "BOULONNERIE"
            return "AUTRES"
            
        data_enriched = []
        for i, item in enumerate(data):
            d = item.copy()
            d['famille'] = get_famille(d.get('designation', ''))
            d['repere'] = item.get('repere', f"Rep-{i+1:03d}")
            data_enriched.append(d)
            
        data_sorted = sorted(data_enriched, key=lambda x: (x['famille'], str(x.get('role', '')), str(x.get('designation', ''))))
        
        current_row = 7
        totals_famille = {}
        total_global = 0.0
        total_surface = 0.0
        
        for famille, group in groupby(data_sorted, key=lambda x: x['famille']):
            ws_details.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
            fam_cell = ws_details.cell(row=current_row, column=1, value=famille)
            fam_cell.font = Font(bold=True, italic=True)
            fam_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            fam_cell.alignment = Alignment(horizontal="left", vertical="center")
            current_row += 1
            
            poids_famille = 0.0
            for item in group:
                qty = item.get('quantity', 1)
                l_mm = float(item.get('length_m', 0)) * 1000 if item.get('length_m') else 0.0
                
                ptot_raw = item.get('poids_total_kg', 0)
                try: ptot = float(ptot_raw)
                except: ptot = 0.0
                
                surf_raw = item.get('surface_peinture_m2', 0)
                try: surf = float(surf_raw)
                except: surf = 0.0
                
                # Compute unit weight if missing
                punt_raw = item.get('poids_unitaire', 0)
                try: punt = float(punt_raw)
                except: punt = 0.0
                
                if punt == 0 and ptot > 0 and qty > 0:
                    punt = ptot / qty
                
                poids_famille += ptot
                total_global += ptot
                total_surface += surf
                
                obs = ""
                if l_mm == 0 and "TÔLES" not in famille and "BOULONNERIE" not in famille:
                    obs = "⚠️ Longueur manquante"
                    
                row_data = [
                    item.get('repere'),
                    str(item.get('role', 'AUTRES')).upper(),
                    str(item.get('designation', '')),
                    "S235", # Nuance default
                    l_mm if l_mm > 0 else "----",
                    qty,
                    round(punt, 2) if punt > 0 else "----",
                    round(ptot, 2),
                    obs
                ]
                
                for col_num, val in enumerate(row_data, 1):
                    c = ws_details.cell(row=current_row, column=col_num, value=val)
                    c.alignment = center_align
                    c.border = thin_border
                    if col_num == 9 and "⚠️" in obs:
                        c.font = Font(color="FF0000")
                current_row += 1
                
            totals_famille[famille] = poids_famille
            
            # Sous-total de famille
            ws_details.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            sub_cell = ws_details.cell(row=current_row, column=1, value=f"Sous-Total {famille}")
            sub_cell.alignment = Alignment(horizontal="right", vertical="center")
            sub_cell.font = Font(bold=True, color="44546A")
            sub_cell.border = thin_border
            
            c_val = ws_details.cell(row=current_row, column=8, value=round(poids_famille, 2))
            c_val.font = Font(bold=True)
            c_val.border = thin_border
            c_val.alignment = center_align
            
            ws_details.cell(row=current_row, column=9).border = thin_border
            
            current_row += 2 # Espace
            
        col_widths_det = {'A': 12, 'B': 25, 'C': 20, 'D': 10, 'E': 15, 'F': 10, 'G': 15, 'H': 15, 'I': 25}
        for col, width in col_widths_det.items():
            ws_details.column_dimensions[col].width = width
            
        # Total Bas de page
        current_row += 1
        ws_details.cell(row=current_row, column=7, value="TOTAL BRUT (Kg)").font = Font(bold=True)
        ws_details.cell(row=current_row, column=8, value=round(total_global, 2)).font = Font(bold=True)
        ws_details.cell(row=current_row, column=8).alignment = center_align
        ws_details.cell(row=current_row, column=7).border = thin_border
        ws_details.cell(row=current_row, column=8).border = thin_border
        current_row += 1
        
        has_boulons = any("BOULON" in str(i.get('designation', '')).upper() or "BOULON" in str(i.get('role', '')).upper() for i in data)
        pourcentage_boulons = 0.02 if has_boulons else 0.05
        poids_boulons = total_global * pourcentage_boulons
        grand_total = total_global + poids_boulons
        
        ws_details.cell(row=current_row, column=7, value="BOULONNERIE/SOUDURE").font = Font(bold=True)
        ws_details.cell(row=current_row, column=8, value=round(poids_boulons, 2)).font = Font(bold=True)
        ws_details.cell(row=current_row, column=8).alignment = center_align
        ws_details.cell(row=current_row, column=7).border = thin_border
        ws_details.cell(row=current_row, column=8).border = thin_border
        current_row += 1
        
        cell_net = ws_details.cell(row=current_row, column=7, value="TOTAL NET (Kg)")
        cell_net.font = Font(bold=True, size=12, color="FFFFFF")
        cell_net.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        cell_net.border = thick_border
        
        cell_net_val = ws_details.cell(row=current_row, column=8, value=round(grand_total, 2))
        cell_net_val.font = Font(bold=True, size=12, color="FFFFFF")
        cell_net_val.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        cell_net_val.border = thick_border
        cell_net_val.alignment = center_align
            
        # --- ONGLET 2: Synthèse Globale ---
        ws_synth = wb.create_sheet(title="Synthèse Globale")
        
        ws_synth.cell(row=2, column=2, value="TABLEAU DE BORD - SYNTHÈSE DU MÉTRÉ").font = Font(bold=True, size=14, color="1F4E78")
        
        synth_headers = ["Catégorie / Famille", "Poids Total (Kg)", "Pourcentage"]
        for col_num, header in enumerate(synth_headers, 2):
            cell = ws_synth.cell(row=5, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            
        r = 6
        for fam, poids in totals_famille.items():
            ws_synth.cell(row=r, column=2, value=fam).border = thin_border
            ws_synth.cell(row=r, column=3, value=round(poids, 2)).border = thin_border
            pct = (poids / grand_total) if grand_total > 0 else 0
            ws_synth.cell(row=r, column=4, value=f"{pct:.1%}").border = thin_border
            ws_synth.cell(row=r, column=4).alignment = center_align
            r += 1
            
        # Add Boulonnerie
        label_boul = "SOUDAGE (2%)" if has_boulons else "BOULONNERIE + SOUDAGE (5%)"
        ws_synth.cell(row=r, column=2, value=label_boul).border = thin_border
        ws_synth.cell(row=r, column=3, value=round(poids_boulons, 2)).border = thin_border
        pct_b = (poids_boulons / grand_total) if grand_total > 0 else 0
        ws_synth.cell(row=r, column=4, value=f"{pct_b:.1%}").border = thin_border
        ws_synth.cell(row=r, column=4).alignment = center_align
        r += 1
        
        # Grand Total
        c_tot_label = ws_synth.cell(row=r, column=2, value="POIDS TOTAL NET")
        c_tot_label.font = Font(bold=True)
        c_tot_label.fill = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
        c_tot_label.border = thin_border
        
        c_tot_val = ws_synth.cell(row=r, column=3, value=round(grand_total, 2))
        c_tot_val.font = Font(bold=True)
        c_tot_val.fill = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
        c_tot_val.border = thin_border
        
        c_tot_pct = ws_synth.cell(row=r, column=4, value="100%")
        c_tot_pct.font = Font(bold=True)
        c_tot_pct.fill = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
        c_tot_pct.border = thin_border
        c_tot_pct.alignment = center_align
        
        r += 3
        ws_synth.cell(row=r, column=2, value="Surface Totale de Peinture (m²):").font = Font(bold=True)
        ws_synth.cell(row=r, column=3, value=round(total_surface, 2)).font = Font(bold=True, color="1F4E78")
        
        col_widths_synth = {'B': 35, 'C': 20, 'D': 15}
        for col, width in col_widths_synth.items():
            ws_synth.column_dimensions[col].width = width

        import io
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()'''

code_new = re.sub(r'    def to_excel_advanced\(.*?\).*?return output\.getvalue\(\)', new_func, code, flags=re.DOTALL)

with open('backend/engines/export_engine.py', 'w', encoding='utf-8') as f:
    f.write(code_new)
