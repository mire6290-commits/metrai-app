import io
import pandas as pd
from typing import List, Dict, Any
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ReportLab imports for PDF generation
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

class ExportEngine:
    @staticmethod
    def to_excel(data: List[Dict[str, Any]]) -> bytes:
        """
        Génère un fichier Excel avancé avec fusion de cellules pour Nomenclatures,
        et calcul final + 5% Boulonnerie.
        """
        from openpyxl import Workbook
        from itertools import groupby

        wb = Workbook()
        ws = wb.active
        ws.title = "Métré Structure"

        # Styles
        header_font = Font(name="Calibri", size=11, bold=True, color="000000")
        header_fill = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        headers = [
            "Pos", "Nomenclatures", "Quantité", "Designation", 
            "Long (mm)", "Poids Kg/(m)", "Poids Kg/Unt", "Poids Tot Kg"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Assurer qu'il y a un champ role pour grouper
        for item in data:
            # The frontend sends 'assemblage' for the role
            role_val = item.get('role') or item.get('assemblage') or item.get('nomenclature') or 'AUTRES'
            item['role'] = str(role_val).upper()

        # Tri pour groupby par role (Nomenclature)
        data_sorted = sorted(data, key=lambda x: str(x.get('role', '')).upper())
        
        current_row = 2
        pos = 1
        total_sum_kg = 0.0
        
        for role_key, group in groupby(data_sorted, key=lambda x: str(x.get('role', '')).upper()):
            group_list = list(group)
            start_row = current_row
            
            for item in group_list:
                ws.cell(row=current_row, column=1, value=pos).alignment = center_align
                ws.cell(row=current_row, column=1).border = thin_border
                
                ws.cell(row=current_row, column=2, value=role_key).alignment = center_align
                ws.cell(row=current_row, column=2).border = thin_border
                
                c_qty = item.get('quantity', item.get('quantite', 0))
                ws.cell(row=current_row, column=3, value=c_qty).alignment = center_align
                ws.cell(row=current_row, column=3).border = thin_border
                
                ws.cell(row=current_row, column=4, value=str(item.get('designation', item.get('profil', '')))).alignment = center_align
                ws.cell(row=current_row, column=4).border = thin_border
                
                val_long = item.get('length_m', item.get('longueur', 0))
                if val_long is not None and val_long > 0:
                    val_long_mm = val_long * 1000 if 'length_m' in item else val_long
                else:
                    val_long_mm = 0
                
                c_long = ws.cell(row=current_row, column=5, value=val_long_mm if val_long_mm else "----")
                c_long.alignment = center_align
                c_long.border = thin_border
                
                val_plin = item.get('masse_lineaire_kg_m', item.get('poids_lineique', 0))
                c_plin = ws.cell(row=current_row, column=6, value=val_plin if val_plin else "----")
                # Formula or static value for Poids Unit
                val_punt_static = item.get('poids_unitaire')
                if val_punt_static is not None and val_punt_static != "----":
                    c_punt = ws.cell(row=current_row, column=7, value=val_punt_static)
                elif val_plin and val_long_mm:
                    f_punt = f"=F{current_row}*(E{current_row}/1000)"
                    c_punt = ws.cell(row=current_row, column=7, value=f_punt)
                else:
                    c_punt = ws.cell(row=current_row, column=7, value="----")
                c_punt.alignment = center_align
                c_punt.border = thin_border
                
                # Formula or static value for Poids Tot Kg
                val_ptot_static = item.get('poids_total_kg', item.get('poids_total'))
                if val_ptot_static is not None and val_ptot_static != "----":
                    c_ptot = ws.cell(row=current_row, column=8, value=val_ptot_static)
                    try:
                        total_sum_kg += float(val_ptot_static)
                    except (ValueError, TypeError):
                        pass
                elif (val_plin and val_long_mm and c_qty) or (val_punt_static is not None and c_qty):
                    f_ptot = f"=G{current_row}*C{current_row}"
                    c_ptot = ws.cell(row=current_row, column=8, value=f_ptot)
                else:
                    c_ptot = ws.cell(row=current_row, column=8, value="----")
                c_ptot.alignment = center_align
                c_ptot.border = thin_border
                
                current_row += 1
                pos += 1
                
            if current_row - 1 > start_row:
                ws.merge_cells(start_row=start_row, start_column=2, end_row=current_row-1, end_column=2)

        # Totaux
        current_row += 1
        
        # Boulonnerie + Soudage 5%
        ws.cell(row=current_row, column=2, value="BOULONNERIE + SOUDAGE").font = Font(bold=True)
        ws.cell(row=current_row, column=2).alignment = center_align
        ws.cell(row=current_row, column=3, value="5%").alignment = center_align
        if current_row > 3:
            ws.cell(row=current_row, column=8, value=round(total_sum_kg * 0.05, 2)).border = thin_border
        
        current_row += 1
        
        # Total Net
        total_cell = ws.cell(row=current_row, column=7, value="Poids Tot Net en Kg")
        total_cell.font = Font(bold=True, italic=True)
        total_cell.fill = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")
        
        if current_row > 4:
            final_sum_cell = ws.cell(row=current_row, column=8, value=round(total_sum_kg * 1.05, 2))
            final_sum_cell.font = Font(bold=True)
            final_sum_cell.fill = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
            final_sum_cell.border = thin_border

        # Ajustement des largeurs
        col_widths = {'A': 5, 'B': 30, 'C': 10, 'D': 25, 'E': 15, 'F': 15, 'G': 15, 'H': 15}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def to_pdf(data: List[Dict[str, Any]]) -> bytes:
        """
        Génère un rapport de métré PDF élégant et structuré.
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=40,
            leftMargin=40,
            topMargin=40,
            bottomMargin=40
        )
        
        story = []
        styles = getSampleStyleSheet()
        
        # Styles personnalisés
        title_style = ParagraphStyle(
            'DocTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=22,
            textColor=colors.HexColor('#1A365D'),
            spaceAfter=15
        )
        
        subtitle_style = ParagraphStyle(
            'DocSubTitle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            textColor=colors.HexColor('#4A5568'),
            spaceAfter=30
        )
        
        th_style = ParagraphStyle(
            'TableHeader',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=9,
            textColor=colors.white,
            alignment=1  # Centré
        )
        
        td_style = ParagraphStyle(
            'TableCell',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=9,
            textColor=colors.HexColor('#2D3748')
        )
        
        td_bold_style = ParagraphStyle(
            'TableCellBold',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=9,
            textColor=colors.HexColor('#1A365D')
        )
        
        # En-tête du document
        story.append(Paragraph("SaaS Métré Charpente", title_style))
        story.append(Paragraph("Rapport technique d'extraction de quantités et métré de structure métallique", subtitle_style))
        story.append(Spacer(1, 10))
        
        # Préparation des données de la table
        table_data = [
            [
                Paragraph("Repère", th_style),
                Paragraph("Profilé", th_style),
                Paragraph("Longueur (mm)", th_style),
                Paragraph("Quantité", th_style),
                Paragraph("Poids Unit. (kg)", th_style),
                Paragraph("Poids Total (kg)", th_style),
                Paragraph("S. Peinture (m²)", th_style)
            ]
        ]
        
        total_poids = 0.0
        total_qty = 0
        total_surf = 0.0
        
        for item in data:
            qty = item.get("quantite", 1)
            poids_t = item.get("poids_total", 0.0)
            surf_t = item.get("surface_total", 0.0)
            
            total_qty += qty
            total_poids += poids_t
            total_surf += surf_t
            
            table_data.append([
                Paragraph(str(item.get("repere", "N/A")), td_style),
                Paragraph(str(item.get("profil", "Inconnu")), td_style),
                Paragraph(f"{item.get('longueur', 0.0):,.0f}", td_style),
                Paragraph(str(qty), td_style),
                Paragraph(f"{item.get('poids_unitaire', 0.0):,.2f}", td_style),
                Paragraph(f"{poids_t:,.2f}", td_style),
                Paragraph(f"{surf_t:,.2f}", td_style)
            ])
            
        # Ligne de Total
        table_data.append([
            Paragraph("TOTAL", td_bold_style),
            Paragraph("", td_style),
            Paragraph("", td_style),
            Paragraph(str(total_qty), td_bold_style),
            Paragraph("", td_style),
            Paragraph(f"{total_poids:,.2f}", td_bold_style),
            Paragraph(f"{total_surf:,.2f}", td_bold_style)
        ])
        
        # Largeurs de colonnes ajustées pour letter size
        col_widths = [60, 90, 80, 50, 80, 85, 85]
        
        t = Table(table_data, colWidths=col_widths)
        t_style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E78')),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-2), 0.5, colors.HexColor('#E2E8F0')),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('TOPPADDING', (0,0), (-1,0), 8),
            # Zebra striping
            ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.white, colors.HexColor('#F8FAFC')]),
            # Total row styling
            ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#EDF2F7')),
            ('LINEABOVE', (0,-1), (-1,-1), 1.5, colors.HexColor('#1F4E78')),
            ('LINEBELOW', (0,-1), (-1,-1), 2, colors.HexColor('#1F4E78')),
            ('TOPPADDING', (0,-1), (-1,-1), 8),
            ('BOTTOMPADDING', (0,-1), (-1,-1), 8),
        ])
        t.setStyle(t_style)
        story.append(t)
        
        # Build Document
        doc.build(story)
        return buffer.getvalue()

    @staticmethod
    def to_csv(data: List[Dict[str, Any]]) -> bytes:
        """
        Génère un export CSV propre.
        """
        df = pd.DataFrame(data)
        output = io.StringIO()
        df.to_csv(output, index=False, encoding='utf-8-sig')
        return output.getvalue().encode('utf-8-sig')
