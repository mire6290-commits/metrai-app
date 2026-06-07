import io
import pandas as pd
from typing import List, Dict, Any
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ReportLab imports for PDF generation
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

class ExportEngine:
    @staticmethod
    def to_excel(data: List[Dict[str, Any]]) -> bytes:
        """
        Génère un fichier Excel avancé avec fusion de cellules pour Nomenclatures,
        et calcul final + 5% Boulonnerie.
        """
        from openpyxl import Workbook
        from itertools import groupby

        wb = Workbook()
        ws = wb.active
        ws.title = "Métré Structure"

        # Styles
        header_font = Font(name="Calibri", size=11, bold=True, color="000000")
        header_fill = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        from datetime import datetime
        
        # En-têtes style Expert
        ws.cell(row=2, column=4, value="Client:").font = Font(bold=True)
        # Default client name or passed project name. For now we use the generic one or try to extract from data
        client_name = "METRAI EXPERT"
        ws.cell(row=2, column=5, value=client_name).font = Font(bold=True)
        
        ws.cell(row=2, column=7, value="Le:").font = Font(bold=True)
        ws.cell(row=2, column=8, value=datetime.now().strftime("%d/%m/%Y")).font = Font(bold=True)
        
        ws.cell(row=4, column=4, value="POIDS NET").font = Font(bold=True, size=12)
        ws.cell(row=4, column=8, value="POIDS BRUT").font = Font(bold=True, size=12)
        
        ws.cell(row=5, column=4, value="1- Ossature métallique :").font = Font(bold=True, underline="single")
        
        headers = [
            "Pos", "Nomenclatures", "Quantité", "Designation", 
            "Long (mm)", "Poids Kg/(m)", "Poids Kg/Unt", "Poids Tot Kg", "S. Peinture (m²)", "Méthode"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Pre-process data to group identical items to avoid duplicate rows
        merged_data = {}
        for item in data:
            role_val = item.get('role') or item.get('assemblage') or item.get('nomenclature') or 'AUTRES'
            role_str = str(role_val).upper()
            
            desig = str(item.get('designation', item.get('profil', ''))).strip()
            long_raw = item.get('length_m', item.get('longueur', 0))
            try:
                l_val = float(long_raw) if long_raw else 0.0
            except:
                l_val = 0.0
            
            # Use unit weight to differentiate if lengths are 0 but unit weights differ
            punt_raw = item.get('poids_unitaire', 0)
            try:
                punt = float(punt_raw) if punt_raw and punt_raw != "----" else 0.0
            except:
                punt = 0.0
                
            qty = int(item.get('quantity', item.get('quantite', 0)))
            ptot_raw = item.get('poids_total_kg', item.get('poids_total', 0))
            try:
                ptot = float(ptot_raw) if ptot_raw and ptot_raw != "----" else 0.0
            except:
                ptot = 0.0
                
            surf_raw = item.get('surface_peinture_m2', 0)
            try:
                surf = float(surf_raw) if surf_raw and surf_raw != "----" else 0.0
            except:
                surf = 0.0
                
            key = f"{role_str}|{desig}|{l_val:.3f}|{punt:.3f}"
            
            if key in merged_data:
                merged_data[key]['quantity'] = merged_data[key].get('quantity', 0) + qty
                if merged_data[key].get('poids_total_kg') is not None and ptot > 0:
                    merged_data[key]['poids_total_kg'] = round(merged_data[key]['poids_total_kg'] + ptot, 2)
                if merged_data[key].get('surface_peinture_m2') is not None and surf > 0:
                    merged_data[key]['surface_peinture_m2'] = round(merged_data[key]['surface_peinture_m2'] + surf, 2)
            else:
                item_copy = item.copy()
                item_copy['role'] = role_str
                item_copy['quantity'] = qty
                if ptot > 0:
                    item_copy['poids_total_kg'] = ptot
                if surf > 0:
                    item_copy['surface_peinture_m2'] = surf
                merged_data[key] = item_copy

        # Tri pour groupby par role (Nomenclature)
        data_sorted = sorted(list(merged_data.values()), key=lambda x: str(x.get('role', '')).upper())
        
        current_row = 8
        pos = 1
        total_sum_kg = 0.0
        
        for role_key, group in groupby(data_sorted, key=lambda x: str(x.get('role', '')).upper()):
            group_list = list(group)
            start_row = current_row
            
            for item in group_list:
                ws.cell(row=current_row, column=1, value=pos).alignment = center_align
                ws.cell(row=current_row, column=1).border = thin_border
                
                ws.cell(row=current_row, column=2, value=role_key).alignment = center_align
                ws.cell(row=current_row, column=2).border = thin_border
                
                c_qty = item.get('quantity', item.get('quantite', 0))
                ws.cell(row=current_row, column=3, value=c_qty).alignment = center_align
                ws.cell(row=current_row, column=3).border = thin_border
                
                ws.cell(row=current_row, column=4, value=str(item.get('designation', item.get('profil', '')))).alignment = center_align
                ws.cell(row=current_row, column=4).border = thin_border
                
                val_long_raw = item.get('length_m', item.get('longueur', 0))
                try:
                    val_long = float(val_long_raw) if val_long_raw else 0.0
                except (ValueError, TypeError):
                    val_long = 0.0
                    
                if val_long > 0:
                    val_long_mm = val_long * 1000 if 'length_m' in item else val_long
                else:
                    val_long_mm = 0
                
                c_long = ws.cell(row=current_row, column=5, value=val_long_mm if val_long_mm else "----")
                c_long.alignment = center_align
                c_long.border = thin_border
                
                val_plin = item.get('masse_lineaire_kg_m', item.get('poids_lineique', 0))
                c_plin = ws.cell(row=current_row, column=6, value=val_plin if val_plin else "----")
                # Formula or static value for Poids Unit
                val_punt_static = item.get('poids_unitaire')
                if val_punt_static is not None and val_punt_static != "----":
                    c_punt = ws.cell(row=current_row, column=7, value=val_punt_static)
                elif val_plin and val_long_mm:
                    f_punt = f"=F{current_row}*(E{current_row}/1000)"
                    c_punt = ws.cell(row=current_row, column=7, value=f_punt)
                else:
                    c_punt = ws.cell(row=current_row, column=7, value="----")
                c_punt.alignment = center_align
                c_punt.border = thin_border
                
                # Formula or static value for Poids Tot Kg
                val_ptot_static = item.get('poids_total_kg', item.get('poids_total'))
                if val_ptot_static is not None and val_ptot_static != "----":
                    c_ptot = ws.cell(row=current_row, column=8, value=val_ptot_static)
                    try:
                        total_sum_kg += float(val_ptot_static)
                    except (ValueError, TypeError):
                        pass
                elif (val_plin and val_long_mm and c_qty) or (val_punt_static is not None and c_qty):
                    f_ptot = f"=G{current_row}*C{current_row}"
                    c_ptot = ws.cell(row=current_row, column=8, value=f_ptot)
                else:
                    c_ptot = ws.cell(row=current_row, column=8, value="----")
                c_ptot.alignment = center_align
                c_ptot.border = thin_border
                
                # Surface Peinture
                val_surf = item.get('surface_peinture_m2')
                c_surf = ws.cell(row=current_row, column=9, value=val_surf if val_surf is not None else "----")
                c_surf.alignment = center_align
                c_surf.border = thin_border
                
                # Méthode
                val_methode = item.get('methode', 'Inconnu')
                c_meth = ws.cell(row=current_row, column=10, value=val_methode)
                c_meth.alignment = center_align
                c_meth.border = thin_border
                
                current_row += 1
                pos += 1
                
            if current_row - 1 > start_row:
                ws.merge_cells(start_row=start_row, start_column=2, end_row=current_row-1, end_column=2)

        last_data_row = current_row - 1

        # Totaux
        current_row += 1
        
        has_bolts = any("BOULON" in str(i.get('role', '')).upper() or "BOULON" in str(i.get('designation', '')).upper() for i in data)

        if has_bolts:
            ws.cell(row=current_row, column=2, value="SOUDAGE (2%)").font = Font(bold=True)
            ws.cell(row=current_row, column=2).alignment = center_align
            ws.cell(row=current_row, column=3, value="2%").alignment = center_align
            boulonnerie_val = total_sum_kg * 0.02
        else:
            ws.cell(row=current_row, column=2, value="BOULONNERIE + SOUDAGE").font = Font(bold=True)
            ws.cell(row=current_row, column=2).alignment = center_align
            ws.cell(row=current_row, column=3, value="5%").alignment = center_align
            boulonnerie_val = total_sum_kg * 0.05
        
        c_boul = ws.cell(row=current_row, column=8, value=round(boulonnerie_val, 3))
        c_boul.border = thin_border
        c_boul.alignment = center_align
        
        current_row += 1
        
        # Total Net
        total_cell = ws.cell(row=current_row, column=7, value="Poids Tot Net en Kg")
        total_cell.font = Font(bold=True, italic=True)
        total_cell.fill = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")
        total_cell.alignment = Alignment(horizontal="right", vertical="center")
        
        total_net_val = total_sum_kg + boulonnerie_val
        final_sum_cell = ws.cell(row=current_row, column=8, value=round(total_net_val, 3))
        final_sum_cell.font = Font(bold=True)
        final_sum_cell.fill = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
        final_sum_cell.border = thin_border
        final_sum_cell.alignment = center_align

        # Ajustement des largeurs
        col_widths = {'A': 5, 'B': 30, 'C': 10, 'D': 25, 'E': 15, 'F': 15, 'G': 20, 'H': 20, 'I': 20, 'J': 15}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def to_excel_advanced(data: List[Dict[str, Any]]) -> bytes:
        """
        Génère un fichier Excel Avancé avec 2 onglets:
        1. Détails de Fabrication (Regroupés par famille)
        2. Synthèse Globale (Dashboard)
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from itertools import groupby
        import re

        wb = Workbook()
        
        # --- ONGLET 1: Détails de Fabrication ---
        ws_details = wb.active
        ws_details.title = "Détails de Fabrication"
        
        # Styles
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        ws_details.cell(row=2, column=2, value="LISTE DE DÉBIT & DÉTAILS").font = Font(bold=True, size=14, color="1F4E78")
        
        headers = ["Repère", "Nomenclature", "Profilé", "Long (mm)", "Quantité", "Poids Unit (Kg)", "Poids Tot (Kg)", "Observation"]
        for col_num, header in enumerate(headers, 1):
            cell = ws_details.cell(row=5, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            
        # Groupement par famille
        def get_famille(designation):
            des = str(designation).upper()
            if re.match(r'^(IPE|HEA|HEB|HEM|UPN|UPE|INP)', des): return "PROFILÉS STANDARDS (I, H, U)"
            if re.match(r'^(L|2L|CORNI)', des): return "CORNIÈRES (L)"
            if re.match(r'^(PL|TN|PLAT|TÔLE)', des): return "TÔLES ET PLATINES"
            if re.match(r'^(TU|CHS|RHS|SHS)', des): return "TUBES"
            if re.match(r'^(BOU|M\d+)', des): return "BOULONNERIE"
            return "AUTRES"
            
        data_enriched = []
        for i, item in enumerate(data):
            d = item.copy()
            d['famille'] = get_famille(d.get('designation', ''))
            d['repere'] = item.get('repere', f"Rep-{i+1:03d}")
            data_enriched.append(d)
            
        data_sorted = sorted(data_enriched, key=lambda x: (x['famille'], str(x.get('role', '')), str(x.get('designation', ''))))
        
        current_row = 6
        totals_famille = {}
        total_global = 0.0
        total_surface = 0.0
        
        for famille, group in groupby(data_sorted, key=lambda x: x['famille']):
            ws_details.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
            fam_cell = ws_details.cell(row=current_row, column=1, value=famille)
            fam_cell.font = Font(bold=True, italic=True)
            fam_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            fam_cell.alignment = Alignment(horizontal="left", vertical="center")
            current_row += 1
            
            poids_famille = 0.0
            for item in group:
                qty = item.get('quantity', 1)
                l_mm = float(item.get('length_m', 0)) * 1000 if item.get('length_m') else 0.0
                
                ptot_raw = item.get('poids_total_kg', 0)
                try: ptot = float(ptot_raw)
                except: ptot = 0.0
                
                surf_raw = item.get('surface_peinture_m2', 0)
                try: surf = float(surf_raw)
                except: surf = 0.0
                
                punt_raw = item.get('poids_unitaire', 0)
                try: punt = float(punt_raw)
                except: punt = 0.0
                
                poids_famille += ptot
                total_global += ptot
                total_surface += surf
                
                obs = ""
                if l_mm == 0 and "TÔLES" not in famille and "BOULONNERIE" not in famille:
                    obs = "⚠️ Longueur manquante"
                    
                row_data = [
                    item.get('repere'),
                    str(item.get('role', 'AUTRES')).upper(),
                    str(item.get('designation', '')),
                    l_mm if l_mm > 0 else "----",
                    qty,
                    round(punt, 2) if punt > 0 else "----",
                    round(ptot, 2),
                    obs
                ]
                
                for col_num, val in enumerate(row_data, 1):
                    c = ws_details.cell(row=current_row, column=col_num, value=val)
                    c.alignment = center_align
                    c.border = thin_border
                    if col_num == 8 and "⚠️" in obs:
                        c.font = Font(color="FF0000")
                current_row += 1
                
            totals_famille[famille] = poids_famille
            current_row += 1 
            
        col_widths_det = {'A': 12, 'B': 25, 'C': 20, 'D': 15, 'E': 10, 'F': 15, 'G': 15, 'H': 25}
        for col, width in col_widths_det.items():
            ws_details.column_dimensions[col].width = width
            
        # --- ONGLET 2: Synthèse Globale ---
        ws_synth = wb.create_sheet(title="Synthèse Globale")
        
        ws_synth.cell(row=2, column=2, value="TABLEAU DE BORD - SYNTHÈSE DU MÉTRÉ").font = Font(bold=True, size=14, color="1F4E78")
        
        synth_headers = ["Catégorie / Famille", "Poids Total (Kg)", "Pourcentage"]
        for col_num, header in enumerate(synth_headers, 2):
            cell = ws_synth.cell(row=5, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            
        r = 6
        has_boulons = any("BOULON" in str(i.get('designation', '')).upper() or "BOULON" in str(i.get('role', '')).upper() for i in data)
        pourcentage_boulons = 0.02 if has_boulons else 0.05
        poids_boulons = total_global * pourcentage_boulons
        
        grand_total = total_global + poids_boulons
        
        for fam, poids in totals_famille.items():
            ws_synth.cell(row=r, column=2, value=fam).border = thin_border
            ws_synth.cell(row=r, column=3, value=round(poids, 2)).border = thin_border
            pct = (poids / grand_total) if grand_total > 0 else 0
            ws_synth.cell(row=r, column=4, value=f"{pct:.1%}").border = thin_border
            ws_synth.cell(row=r, column=4).alignment = center_align
            r += 1
            
        # Add Boulonnerie
        label_boul = "SOUDAGE (2%)" if has_boulons else "BOULONNERIE + SOUDAGE (5%)"
        ws_synth.cell(row=r, column=2, value=label_boul).border = thin_border
        ws_synth.cell(row=r, column=3, value=round(poids_boulons, 2)).border = thin_border
        pct_b = (poids_boulons / grand_total) if grand_total > 0 else 0
        ws_synth.cell(row=r, column=4, value=f"{pct_b:.1%}").border = thin_border
        ws_synth.cell(row=r, column=4).alignment = center_align
        r += 1
        
        # Grand Total
        c_tot_label = ws_synth.cell(row=r, column=2, value="POIDS TOTAL NET")
        c_tot_label.font = Font(bold=True)
        c_tot_label.fill = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
        c_tot_label.border = thin_border
        
        c_tot_val = ws_synth.cell(row=r, column=3, value=round(grand_total, 2))
        c_tot_val.font = Font(bold=True)
        c_tot_val.fill = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
        c_tot_val.border = thin_border
        
        c_tot_pct = ws_synth.cell(row=r, column=4, value="100%")
        c_tot_pct.font = Font(bold=True)
        c_tot_pct.fill = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
        c_tot_pct.border = thin_border
        c_tot_pct.alignment = center_align
        
        r += 3
        ws_synth.cell(row=r, column=2, value="Surface Totale de Peinture (m²):").font = Font(bold=True)
        ws_synth.cell(row=r, column=3, value=round(total_surface, 2)).font = Font(bold=True, color="1F4E78")
        
        col_widths_synth = {'B': 35, 'C': 20, 'D': 15}
        for col, width in col_widths_synth.items():
            ws_synth.column_dimensions[col].width = width

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def to_pdf(data: List[Dict[str, Any]]) -> bytes:
        """
        Génère un rapport de métré PDF élégant et structuré.
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=40,
            leftMargin=40,
            topMargin=40,
            bottomMargin=40
        )
        
        story = []
        styles = getSampleStyleSheet()
        
        # Styles personnalisés
        title_style = ParagraphStyle(
            'DocTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=22,
            textColor=colors.HexColor('#1A365D'),
            spaceAfter=15
        )
        
        subtitle_style = ParagraphStyle(
            'DocSubTitle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            textColor=colors.HexColor('#4A5568'),
            spaceAfter=30
        )
        
        th_style = ParagraphStyle(
            'TableHeader',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=9,
            textColor=colors.white,
            alignment=1  # Centré
        )
        
        td_style = ParagraphStyle(
            'TableCell',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=9,
            textColor=colors.HexColor('#2D3748')
        )
        
        td_bold_style = ParagraphStyle(
            'TableCellBold',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=9,
            textColor=colors.HexColor('#1A365D')
        )
        
        # En-tête du document
        story.append(Paragraph("SaaS Métré Charpente", title_style))
        story.append(Paragraph("Rapport technique d'extraction de quantités et métré de structure métallique", subtitle_style))
        story.append(Spacer(1, 10))
        
        # Préparation des données de la table
        table_data = [
            [
                Paragraph("Repère", th_style),
                Paragraph("Profilé", th_style),
                Paragraph("Longueur (mm)", th_style),
                Paragraph("Quantité", th_style),
                Paragraph("Poids Unit. (kg)", th_style),
                Paragraph("Poids Total (kg)", th_style),
                Paragraph("S. Peinture (m²)", th_style)
            ]
        ]
        
        total_poids = 0.0
        total_qty = 0
        total_surf = 0.0
        
        for item in data:
            qty = item.get("quantite", 1)
            poids_t = item.get("poids_total", 0.0)
            surf_t = item.get("surface_total", 0.0)
            
            total_qty += qty
            total_poids += poids_t
            total_surf += surf_t
            
            try:
                longueur_float = float(item.get('longueur', 0.0) or 0.0)
            except (ValueError, TypeError):
                longueur_float = 0.0

            table_data.append([
                Paragraph(str(item.get("repere", "N/A")), td_style),
                Paragraph(str(item.get("profil", "Inconnu")), td_style),
                Paragraph(f"{longueur_float:,.0f}", td_style),
                Paragraph(str(qty), td_style),
                Paragraph(f"{item.get('poids_unitaire', 0.0):,.2f}", td_style),
                Paragraph(f"{poids_t:,.2f}", td_style),
                Paragraph(f"{surf_t:,.2f}", td_style)
            ])
            
        # Ligne de Total
        table_data.append([
            Paragraph("TOTAL", td_bold_style),
            Paragraph("", td_style),
            Paragraph("", td_style),
            Paragraph(str(total_qty), td_bold_style),
            Paragraph("", td_style),
            Paragraph(f"{total_poids:,.2f}", td_bold_style),
            Paragraph(f"{total_surf:,.2f}", td_bold_style)
        ])
        
        # Largeurs de colonnes ajustées pour letter size
        col_widths = [60, 90, 80, 50, 80, 85, 85]
        
        t = Table(table_data, colWidths=col_widths)
        t_style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E78')),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-2), 0.5, colors.HexColor('#E2E8F0')),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('TOPPADDING', (0,0), (-1,0), 8),
            # Zebra striping
            ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.white, colors.HexColor('#F8FAFC')]),
            # Total row styling
            ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#EDF2F7')),
            ('LINEABOVE', (0,-1), (-1,-1), 1.5, colors.HexColor('#1F4E78')),
            ('LINEBELOW', (0,-1), (-1,-1), 2, colors.HexColor('#1F4E78')),
            ('TOPPADDING', (0,-1), (-1,-1), 8),
            ('BOTTOMPADDING', (0,-1), (-1,-1), 8),
        ])
        t.setStyle(t_style)
        story.append(t)
        
        # Build Document
        doc.build(story)
        return buffer.getvalue()

    @staticmethod
    def to_csv(data: List[Dict[str, Any]]) -> bytes:
        """
        Génère un export CSV propre.
        """
        df = pd.DataFrame(data)
        output = io.StringIO()
        df.to_csv(output, index=False, encoding='utf-8-sig')
        return output.getvalue().encode('utf-8-sig')
